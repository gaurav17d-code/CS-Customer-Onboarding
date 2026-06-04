import json
import os
import shutil
from datetime import datetime
from database import get_resolutions, get_case

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

CONFIG_DIR = os.environ.get('CONFIG_DIR', 'config')
CMR_TEMPLATE_PATH = os.environ.get('CMR_TEMPLATE', 'templates/CMR_template.xlsx')
OUTPUT_DIR = os.environ.get('OUTPUT_DIR', 'storage/cmr_output')

def load_mapping():
    path = os.path.join(CONFIG_DIR, 'cmr_mapping.json')
    with open(path) as f:
        return json.load(f)

def ensure_output_dir():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

def generate_cmr(case_id: str, manual_fields: dict = None) -> dict:
    """
    Generate CMR Excel workbook from approved field resolutions.
    manual_fields: dict of {canonical_name: value} for fields entered manually.
    Returns {'path': str, 'unresolved': list, 'status': str}
    """
    ensure_output_dir()

    if not OPENPYXL_AVAILABLE:
        return {'path': None, 'unresolved': [], 'status': 'error',
                'message': 'openpyxl not installed'}

    # Load mapping config
    try:
        mapping = load_mapping()
    except Exception as e:
        return {'path': None, 'unresolved': [], 'status': 'error',
                'message': f'Cannot load CMR mapping: {e}'}

    field_mappings = mapping.get('field_mappings', [])

    # Get approved resolutions from DB
    resolutions = get_resolutions(case_id)
    case = get_case(case_id)

    if manual_fields is None:
        manual_fields = {}

    # Build final values dict
    final_values = {}
    unresolved = []

    for fm in field_mappings:
        canonical = fm['canonical']
        cmr_label = fm['cmr_label']
        sources = fm.get('sources', [])

        value = None

        # 1. Check manual fields first
        if canonical in manual_fields and manual_fields[canonical]:
            value = manual_fields[canonical]

        # 2. Check reviewer resolutions
        elif canonical in resolutions:
            value = resolutions[canonical].get('final_value')

        # 3. Check case metadata fields
        elif 'case_metadata' in sources and case:
            meta_map = {
                'sales_officer': 'sales_officer',
                'sales_area': 'sales_area',
                'product': 'product_type',
                'regional_office': 'regional_office'
            }
            meta_key = meta_map.get(canonical)
            if meta_key:
                value = case.get(meta_key)

        if value:
            final_values[canonical] = value
        else:
            unresolved.append({'canonical': canonical, 'label': cmr_label})

    # Check if template exists
    if not os.path.exists(CMR_TEMPLATE_PATH):
        # Create a basic workbook if template doesn't exist
        return _generate_basic_cmr(case_id, final_values, unresolved, field_mappings, case)

    # Load and populate the template
    try:
        wb = load_workbook(CMR_TEMPLATE_PATH)
        ws = wb.active

        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        red_font = Font(color='FF0000', bold=True)

        for fm in field_mappings:
            canonical = fm['canonical']
            cell_ref = fm.get('cmr_cell')
            if not cell_ref:
                continue

            if canonical in final_values:
                ws[cell_ref] = final_values[canonical]
            else:
                ws[cell_ref] = 'PENDING'
                cell = ws[cell_ref]
                cell.fill = yellow_fill
                cell.font = red_font

        output_filename = f'CMR_{case_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        output_path = os.path.join(OUTPUT_DIR, output_filename)
        wb.save(output_path)

        status = 'draft' if unresolved else 'complete'
        return {
            'path': output_path,
            'filename': output_filename,
            'unresolved': unresolved,
            'final_values': final_values,
            'status': status
        }

    except Exception as e:
        return _generate_basic_cmr(case_id, final_values, unresolved, field_mappings, case)

def _generate_basic_cmr(case_id: str, final_values: dict, unresolved: list,
                         field_mappings: list, case: dict) -> dict:
    """Generate a basic CMR workbook without a template."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = 'CMR'

        # Header
        ws['A1'] = 'Customer Master Record (CMR)'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f'Case ID: {case_id}'
        ws['A3'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}, Status: {"DRAFT - Fields Pending" if unresolved else "COMPLETE"}'
        ws['A3'].font = Font(italic=True)

        ws['A5'] = 'Field'
        ws['B5'] = 'Value'
        ws['C5'] = 'Status'
        for col in ['A5', 'B5', 'C5']:
            ws[col].font = Font(bold=True)

        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

        row = 6
        for fm in field_mappings:
            canonical = fm['canonical']
            label = fm['cmr_label']
            ws.cell(row=row, column=1, value=label)
            if canonical in final_values:
                ws.cell(row=row, column=2, value=final_values[canonical])
                ws.cell(row=row, column=3, value='Approved')
                ws.cell(row=row, column=2).fill = green_fill
            else:
                ws.cell(row=row, column=2, value='PENDING')
                ws.cell(row=row, column=3, value='Unresolved')
                ws.cell(row=row, column=2).fill = yellow_fill
            row += 1

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 15

        ensure_output_dir()
        output_filename = f'CMR_{case_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        output_path = os.path.join(OUTPUT_DIR, output_filename)
        wb.save(output_path)

        status = 'draft' if unresolved else 'complete'
        return {
            'path': output_path,
            'filename': output_filename,
            'unresolved': unresolved,
            'final_values': final_values,
            'status': status
        }
    except Exception as e:
        return {'path': None, 'unresolved': unresolved, 'status': 'error',
                'message': str(e)}
